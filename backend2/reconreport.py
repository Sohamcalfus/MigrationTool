from zeep import Client
from zeep.wsse.username import UsernameToken
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
from datetime import datetime
import tempfile
import uuid

# SOAP API Configuration (keeping your original config)
SOAP_CONFIG = {
    'wsdl_url': 'https://miterbrands-ibayqy-test.fa.ocs.oraclecloud.com/xmlpserver/services/v2/ReportService?WSDL',
    'username': 'FUSTST.CONVERSION',
    'password': 'Conversion@2025',
    'target_report_path': '/Custom/Migration Reports/Report/CUSTOMER_RECON_REPORT.xdo'
}

# Only raw file path needed (keeping your original)
RAW_FILE_PATH = fr"C:\Project\MigrationTool\Customer Invoice Conversion Data - RAW File.xlsx"

# Keep your original functions exactly as they are
def fetch_target_data_via_soap():
    """Fetch target report data using SOAP API"""
    try:
        print("🔗 Connecting to SOAP API...")
        client = Client(wsdl=SOAP_CONFIG['wsdl_url'], 
                       wsse=UsernameToken(SOAP_CONFIG['username'], SOAP_CONFIG['password']))
        
        # SOAP parameters with all required fields
        report_request = {
            'reportAbsolutePath': SOAP_CONFIG['target_report_path'],
            'sizeOfDataChunkDownload': '-1',
            'byPassCache': True,
            'flattenXML': False,
            'parameterNameValues': {
                'listOfParamNameValues': {
                    'item': [
                        {
                            'name': 'P_ORG_ID',
                            'values': {'item': ['300000003170678']},
                            'templateParam': False,
                            'multiValuesAllowed': False,
                            'refreshParamOnChange': False,
                            'selectAll': False,
                            'useNullForAll': False
                        },
                        {
                            'name': 'P_FROM_DATE',
                            'values': {'item': ['2025-01-01']},
                            'templateParam': False,
                            'multiValuesAllowed': False,
                            'refreshParamOnChange': False,
                            'selectAll': False,
                            'useNullForAll': False
                        },
                        {
                            'name': 'P_TO_DATE',
                            'values': {'item': ['2025-12-31']},
                            'templateParam': False,
                            'multiValuesAllowed': False,
                            'refreshParamOnChange': False,
                            'selectAll': False,
                            'useNullForAll': False
                        },
                        {
                            'name': 'P_CURRENCY_CODE',
                            'values': {'item': ['USD']},
                            'templateParam': False,
                            'multiValuesAllowed': False,
                            'refreshParamOnChange': False,
                            'selectAll': False,
                            'useNullForAll': False
                        }
                    ]
                }
            }
        }
        
        print("📤 Executing SOAP request...")
        response = client.service.runReport(report_request, SOAP_CONFIG['username'], SOAP_CONFIG['password'])
        
        # Save temporary file
        temp_file = f"temp_target_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xls"
        with open(temp_file, "wb") as f:
            f.write(response.reportBytes)
        
        # Read with pandas
        target_df = pd.read_excel(temp_file, engine='xlrd')
        os.remove(temp_file)  # Cleanup
        
        print(f"✅ Target data fetched: {len(target_df)} records, {len(target_df.columns)} columns")
        print("REPORT file columns:", target_df.columns.tolist())
        return target_df
        
    except Exception as e:
        print(f"❌ SOAP Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def generate_reconciliation_report():
    """Main function following your exact logic structure"""
    try:
        print("🚀 Starting Reconciliation Report Generation...")
        
        # Auto-generate output path
        raw_dir = os.path.dirname(RAW_FILE_PATH)
        raw_name = os.path.splitext(os.path.basename(RAW_FILE_PATH))[0]
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = os.path.join(raw_dir, f"{raw_name}_Reconciliation_{timestamp}.xlsx")
        
        # Step 1: Fetch target data via SOAP
        target_df = fetch_target_data_via_soap()
        if target_df is None:
            raise Exception("Failed to fetch target data via SOAP API")
        
        # Step 2: Read source data (SKIP FIRST ROW like your code)
        print("📊 Reading source data...")
        raw_df = pd.read_excel(RAW_FILE_PATH, skiprows=1)
        
        # Clean column names
        raw_df.columns = raw_df.columns.str.strip()
        target_df.columns = target_df.columns.str.strip()
        
        print("RAW file columns:", raw_df.columns.tolist())
        print("REPORT file columns:", target_df.columns.tolist())
        print(f"RAW file records: {len(raw_df)}")
        print(f"REPORT file records: {len(target_df)}")
        
        # Step 3: Verify exact column names exist (like your code)
        required_raw_columns = [
            "Bill-to Customer Account Number",
            "Transaction Number", 
            "Transaction Line Amount",
            "Line Transactions Flexfield Segment 3"  # KEPT AS ORIGINAL
        ]
        
        required_report_columns = [
            "Customer Number",
            "Invoice Number",
            "Amount (Header & Line)",
            "Line Attribute 3"  # CHANGED FROM "Line Transactions Flexfield Segment 3"
        ]
        
        print("\nChecking RAW columns...")
        for col in required_raw_columns:
            if col in raw_df.columns:
                print(f"✅ Found: '{col}'")
            else:
                print(f"❌ Missing: '{col}'")
                print(f"Available: {[c for c in raw_df.columns if 'customer' in c.lower() or 'transaction' in c.lower() or 'amount' in c.lower() or 'segment' in c.lower()]}")
        
        print("\nChecking REPORT columns...")
        for col in required_report_columns:
            if col in target_df.columns:
                print(f"✅ Found: '{col}'")
            else:
                print(f"❌ Missing: '{col}'")
                print(f"Available: {[c for c in target_df.columns if 'customer' in c.lower() or 'invoice' in c.lower() or 'amount' in c.lower() or 'segment' in c.lower()]}")
        
        # Step 4: Convert to string for safe comparison (except Amount - handle as integer)
        print("\nPreparing data...")
        
        # RAW file key columns - KEPT ORIGINAL COLUMN NAME
        raw_df["Bill-to Customer Account Number"] = raw_df["Bill-to Customer Account Number"].fillna("").astype(str).str.strip()
        raw_df["Transaction Number"] = raw_df["Transaction Number"].fillna("").astype(str).str.strip()
        # Convert Amount to integer (not string)
        raw_df["Transaction Line Amount"] = pd.to_numeric(raw_df["Transaction Line Amount"], errors='coerce').fillna(0).astype(int)
        raw_df["Line Transactions Flexfield Segment 3"] = raw_df["Line Transactions Flexfield Segment 3"].fillna("").astype(str).str.strip()  # KEPT ORIGINAL
        
        # REPORT file key columns - CHANGED TO "Line Attribute 3"
        target_df["Customer Number"] = target_df["Customer Number"].fillna("").astype(str).str.strip()
        target_df["Invoice Number"] = target_df["Invoice Number"].fillna("").astype(str).str.strip()
        # Convert Amount to integer (not string)
        target_df["Amount (Header & Line)"] = pd.to_numeric(target_df["Amount (Header & Line)"], errors='coerce').fillna(0).astype(int)
        target_df["Line Attribute 3"] = target_df["Line Attribute 3"].fillna("").astype(str).str.strip()  # CHANGED TO "Line Attribute 3"
        
        # Step 5: Cross-comparison logic (EXACTLY like your code)
        output_rows = []
        
        print("Starting cross-comparison...")
        for raw_idx, raw_row in raw_df.iterrows():
            if (raw_idx + 1) % 50 == 0:
                print(f"Processing record {raw_idx + 1}/{len(raw_df)}")
            
            # Extract RAW values - KEPT ORIGINAL COLUMN NAME
            raw_customer = raw_row["Bill-to Customer Account Number"]
            raw_invoice = raw_row["Transaction Number"]
            raw_amount = raw_row["Transaction Line Amount"]
            raw_segment = raw_row["Line Transactions Flexfield Segment 3"]  # KEPT ORIGINAL
            
            # Find matching REPORT record by customer + invoice
            matching_reports = target_df[
                (target_df["Customer Number"] == raw_customer) &
                (target_df["Invoice Number"] == raw_invoice)
            ]
            
            # Build output row
            output_row = {}
            
            # SECTION 1: RAW VALUES (LEFT SIDE) - KEPT ORIGINAL COLUMN NAME
            output_row["RAW: Bill-to Customer Account Number"] = raw_customer
            output_row["RAW: Transaction Number"] = raw_invoice
            output_row["RAW: Transaction Line Amount"] = raw_amount
            output_row["RAW: Line Transactions Flexfield Segment 3"] = raw_segment  # KEPT ORIGINAL
            
            if len(matching_reports) > 0:
                # Match found
                report_row = matching_reports.iloc[0]
                
                report_customer = report_row["Customer Number"]
                report_invoice = report_row["Invoice Number"]
                report_amount = report_row["Amount (Header & Line)"]
                report_segment = report_row["Line Attribute 3"]  # CHANGED TO "Line Attribute 3"
                
                # SECTION 2: MATCH STATUS (MIDDLE) - KEPT ORIGINAL COLUMN NAME FOR RAW SIDE
                output_row["MATCH: Bill-to Customer Account Number"] = (raw_customer == report_customer) and raw_customer != ""
                output_row["MATCH: Transaction Number"] = (raw_invoice == report_invoice) and raw_invoice != ""
                output_row["MATCH: Transaction Line Amount"] = (raw_amount == report_amount) and raw_amount != 0 and report_amount != 0
                output_row["MATCH: Line Transactions Flexfield Segment 3"] = (raw_segment == report_segment) and raw_segment != "" and report_segment != ""  # KEPT ORIGINAL
                
                # SECTION 3: REPORT VALUES (RIGHT SIDE) - CHANGED TO "Line Attribute 3"
                output_row["REPORT: Customer Number"] = report_customer
                output_row["REPORT: Invoice Number"] = report_invoice
                output_row["REPORT: Amount (Header & Line)"] = report_amount
                output_row["REPORT: Line Attribute 3"] = report_segment  # CHANGED TO "Line Attribute 3"
                
            else:
                # No match found
                output_row["MATCH: Bill-to Customer Account Number"] = False
                output_row["MATCH: Transaction Number"] = False
                output_row["MATCH: Transaction Line Amount"] = False
                output_row["MATCH: Line Transactions Flexfield Segment 3"] = False  # KEPT ORIGINAL
                
                output_row["REPORT: Customer Number"] = "NO MATCH FOUND"
                output_row["REPORT: Invoice Number"] = "NO MATCH FOUND"
                output_row["REPORT: Amount (Header & Line)"] = "NO MATCH FOUND"
                output_row["REPORT: Line Attribute 3"] = "NO MATCH FOUND"  # CHANGED TO "Line Attribute 3"
            
            output_rows.append(output_row)
        
        # Step 6: Create final DataFrame
        output_df = pd.DataFrame(output_rows)
        
        print(f"Generated {len(output_df)} comparison rows")
        
        # Step 7: Write to Excel
        print("Writing to Excel...")
        output_df.to_excel(output_path, index=False, engine='openpyxl')
        
        # Step 8: Apply colors to MATCH columns (EXACTLY like your code)
        print("Applying colors...")
        wb = load_workbook(output_path)
        ws = wb.active
        
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Find MATCH columns and color them
        for col_idx in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=1, column=col_idx)
            if header_cell.value and str(header_cell.value).startswith("MATCH:"):
                print(f"Coloring column {col_idx}: {header_cell.value}")
                
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if str(cell.value).strip().upper() == "TRUE":
                        cell.fill = green_fill
                    elif str(cell.value).strip().upper() == "FALSE":
                        cell.fill = red_fill
        
        wb.save(output_path)
        
        print(f"\n✅ SUCCESS!")
        print(f"Reconciliation report saved to: {output_path}")
        print(f"Total records processed: {len(output_df)}")
        
        # Summary statistics (EXACTLY like your code)
        matched_records = sum(1 for row in output_rows
                             if row["MATCH: Bill-to Customer Account Number"] and row["MATCH: Transaction Number"])
        print(f"Records with matching Customer + Invoice: {matched_records}")
        
        return {
            "status": "success",
            "output_file": output_path,
            "total_records": len(output_df),
            "matched_records": matched_records,
            "source_file": RAW_FILE_PATH
        }
        
    except Exception as e:
        print(f"❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return {"status": "error", "error": str(e)}

# ADD THIS CLASS FOR FLASK INTEGRATION
class ReconciliationReportGenerator:
    def __init__(self, soap_config):
        self.soap_config = soap_config
        
    def fetch_target_data_via_soap(self):
        """Fetch target report data using SOAP API"""
        try:
            print("🔗 Connecting to SOAP API...")
            client = Client(wsdl=self.soap_config['wsdl_url'], 
                           wsse=UsernameToken(self.soap_config['username'], self.soap_config['password']))
            
            # SOAP parameters with all required fields
            report_request = {
                'reportAbsolutePath': self.soap_config['target_report_path'],
                'sizeOfDataChunkDownload': '-1',
                'byPassCache': True,
                'flattenXML': False,
                'parameterNameValues': {
                    'listOfParamNameValues': {
                        'item': [
                            {
                                'name': 'P_ORG_ID',
                                'values': {'item': ['300000003170678']},
                                'templateParam': False,
                                'multiValuesAllowed': False,
                                'refreshParamOnChange': False,
                                'selectAll': False,
                                'useNullForAll': False
                            },
                            {
                                'name': 'P_FROM_DATE',
                                'values': {'item': ['2025-01-01']},
                                'templateParam': False,
                                'multiValuesAllowed': False,
                                'refreshParamOnChange': False,
                                'selectAll': False,
                                'useNullForAll': False
                            },
                            {
                                'name': 'P_TO_DATE',
                                'values': {'item': ['2025-12-31']},
                                'templateParam': False,
                                'multiValuesAllowed': False,
                                'refreshParamOnChange': False,
                                'selectAll': False,
                                'useNullForAll': False
                            },
                            {
                                'name': 'P_CURRENCY_CODE',
                                'values': {'item': ['USD']},
                                'templateParam': False,
                                'multiValuesAllowed': False,
                                'refreshParamOnChange': False,
                                'selectAll': False,
                                'useNullForAll': False
                            }
                        ]
                    }
                }
            }
            
            print("📤 Executing SOAP request...")
            response = client.service.runReport(report_request, self.soap_config['username'], self.soap_config['password'])
            
            # Save temporary file
            temp_file = f"temp_target_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xls"
            with open(temp_file, "wb") as f:
                f.write(response.reportBytes)
            
            # Read with pandas
            target_df = pd.read_excel(temp_file, engine='xlrd')
            os.remove(temp_file)  # Cleanup
            
            print(f"✅ Target data fetched: {len(target_df)} records, {len(target_df.columns)} columns")
            return target_df
            
        except Exception as e:
            print(f"❌ SOAP Error: {str(e)}")
            raise Exception(f"Failed to fetch target data: {str(e)}")

    def generate_reconciliation_report(self, raw_file_path, output_dir=None):
        """Generate reconciliation report for Flask API"""
        try:
            print("🚀 Starting Reconciliation Report Generation...")
            
            # Create output directory if not provided
            if output_dir is None:
                output_dir = tempfile.gettempdir()
            
            # Generate unique output filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            unique_id = str(uuid.uuid4())[:8]
            output_filename = f"Reconciliation_Report_{timestamp}_{unique_id}.xlsx"
            output_path = os.path.join(output_dir, output_filename)
            
            # Step 1: Fetch target data via SOAP
            target_df = self.fetch_target_data_via_soap()
            if target_df is None:
                raise Exception("Failed to fetch target data via SOAP API")
            
            # Step 2: Read source data
            print("📊 Reading source data...")
            raw_df = pd.read_excel(raw_file_path, skiprows=1)
            
            # Clean column names
            raw_df.columns = raw_df.columns.str.strip()
            target_df.columns = target_df.columns.str.strip()
            
            print(f"RAW file records: {len(raw_df)}")
            print(f"REPORT file records: {len(target_df)}")
            
            # Step 3: Verify required columns
            required_raw_columns = [
                "Bill-to Customer Account Number",
                "Transaction Number", 
                "Transaction Line Amount",
                "Line Transactions Flexfield Segment 3"
            ]
            
            required_report_columns = [
                "Customer Number",
                "Invoice Number",
                "Amount (Header & Line)",
                "Line Attribute 3"
            ]
            
            # Check if all required columns exist
            missing_raw_cols = [col for col in required_raw_columns if col not in raw_df.columns]
            missing_report_cols = [col for col in required_report_columns if col not in target_df.columns]
            
            if missing_raw_cols:
                raise Exception(f"Missing RAW columns: {missing_raw_cols}")
            if missing_report_cols:
                raise Exception(f"Missing REPORT columns: {missing_report_cols}")
            
            # Step 4: Data preparation
            print("Preparing data...")
            
            # RAW file key columns
            raw_df["Bill-to Customer Account Number"] = raw_df["Bill-to Customer Account Number"].fillna("").astype(str).str.strip()
            raw_df["Transaction Number"] = raw_df["Transaction Number"].fillna("").astype(str).str.strip()
            raw_df["Transaction Line Amount"] = pd.to_numeric(raw_df["Transaction Line Amount"], errors='coerce').fillna(0).astype(int)
            raw_df["Line Transactions Flexfield Segment 3"] = raw_df["Line Transactions Flexfield Segment 3"].fillna("").astype(str).str.strip()
            
            # REPORT file key columns
            target_df["Customer Number"] = target_df["Customer Number"].fillna("").astype(str).str.strip()
            target_df["Invoice Number"] = target_df["Invoice Number"].fillna("").astype(str).str.strip()
            target_df["Amount (Header & Line)"] = pd.to_numeric(target_df["Amount (Header & Line)"], errors='coerce').fillna(0).astype(int)
            target_df["Line Attribute 3"] = target_df["Line Attribute 3"].fillna("").astype(str).str.strip()
            
            # Step 5: Cross-comparison logic
            output_rows = []
            total_records = len(raw_df)
            
            print("Starting cross-comparison...")
            for raw_idx, raw_row in raw_df.iterrows():
                if (raw_idx + 1) % 50 == 0:
                    print(f"Processing record {raw_idx + 1}/{total_records}")
                
                # Extract RAW values
                raw_customer = raw_row["Bill-to Customer Account Number"]
                raw_invoice = raw_row["Transaction Number"]
                raw_amount = raw_row["Transaction Line Amount"]
                raw_segment = raw_row["Line Transactions Flexfield Segment 3"]
                
                # Find matching REPORT record
                matching_reports = target_df[
                    (target_df["Customer Number"] == raw_customer) &
                    (target_df["Invoice Number"] == raw_invoice)
                ]
                
                # Build output row
                output_row = {
                    "RAW: Bill-to Customer Account Number": raw_customer,
                    "RAW: Transaction Number": raw_invoice,
                    "RAW: Transaction Line Amount": raw_amount,
                    "RAW: Line Transactions Flexfield Segment 3": raw_segment
                }
                
                if len(matching_reports) > 0:
                    # Match found
                    report_row = matching_reports.iloc[0]
                    
                    report_customer = report_row["Customer Number"]
                    report_invoice = report_row["Invoice Number"]
                    report_amount = report_row["Amount (Header & Line)"]
                    report_segment = report_row["Line Attribute 3"]
                    
                    # Match status
                    output_row.update({
                        "MATCH: Bill-to Customer Account Number": (raw_customer == report_customer) and raw_customer != "",
                        "MATCH: Transaction Number": (raw_invoice == report_invoice) and raw_invoice != "",
                        "MATCH: Transaction Line Amount": (raw_amount == report_amount) and raw_amount != 0 and report_amount != 0,
                        "MATCH: Line Transactions Flexfield Segment 3": (raw_segment == report_segment) and raw_segment != "" and report_segment != "",
                        "REPORT: Customer Number": report_customer,
                        "REPORT: Invoice Number": report_invoice,
                        "REPORT: Amount (Header & Line)": report_amount,
                        "REPORT: Line Attribute 3": report_segment
                    })
                else:
                    # No match found
                    output_row.update({
                        "MATCH: Bill-to Customer Account Number": False,
                        "MATCH: Transaction Number": False,
                        "MATCH: Transaction Line Amount": False,
                        "MATCH: Line Transactions Flexfield Segment 3": False,
                        "REPORT: Customer Number": "NO MATCH FOUND",
                        "REPORT: Invoice Number": "NO MATCH FOUND",
                        "REPORT: Amount (Header & Line)": "NO MATCH FOUND",
                        "REPORT: Line Attribute 3": "NO MATCH FOUND"
                    })
                
                output_rows.append(output_row)
            
            # Step 6: Create final DataFrame and save
            output_df = pd.DataFrame(output_rows)
            print("Writing to Excel...")
            output_df.to_excel(output_path, index=False, engine='openpyxl')
            
            # Step 7: Apply colors
            print("Applying colors...")
            wb = load_workbook(output_path)
            ws = wb.active
            
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            for col_idx in range(1, ws.max_column + 1):
                header_cell = ws.cell(row=1, column=col_idx)
                if header_cell.value and str(header_cell.value).startswith("MATCH:"):
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if str(cell.value).strip().upper() == "TRUE":
                            cell.fill = green_fill
                        elif str(cell.value).strip().upper() == "FALSE":
                            cell.fill = red_fill
            
            wb.save(output_path)
            
            # Calculate statistics
            matched_records = sum(1 for row in output_rows
                                 if row["MATCH: Bill-to Customer Account Number"] and row["MATCH: Transaction Number"])
            
            print(f"✅ SUCCESS! Reconciliation report saved to: {output_path}")
            
            return {
                "status": "success",
                "output_file": output_path,
                "output_filename": output_filename,
                "total_records": len(output_df),
                "matched_records": matched_records,
                "match_percentage": round((matched_records / len(output_df)) * 100, 2) if len(output_df) > 0 else 0
            }
            
        except Exception as e:
            print(f"❌ ERROR: {str(e)}")
            return {"status": "error", "error": str(e)}

# Keep your original standalone execution
if __name__ == "__main__":
    result = generate_reconciliation_report()
    if result["status"] == "success":
        print(f"\n🎉 Reconciliation completed!")
        print(f"📂 Check: {result['output_file']}")
    else:
        print(f"\n💥 Failed: {result['error']}")
